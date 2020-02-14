import string
import random
import uuid
import base64, re
import binascii
import aiohttp
import copy
from gatco.response import json, text, html, file
from application.extensions import apimanager
from application.database import db
from application.models.models import *


from application.server import app
from gatco_apimanager.views.sqlalchemy.helpers import to_dict

from sqlalchemy.sql.expression import except_
import time
from datetime import datetime
from math import floor
from application.client import HTTPClient
from sqlalchemy import or_, and_, desc
from application.extensions import auth
import ujson
import xlrender
# import xlrenderpy as xlrender
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

from gatco import Blueprint
exportbp = Blueprint('exportbp',url_prefix='/export')

@exportbp.route('/me')
async def me(request):
    return json({'my': 'blueprint'})

# @app.route('/exportexcel/test', methods=['GET'])
# async def get_token(request):
#     filename = "test.xlsx"
#     deffile = "sample.json"
#     template = "sample.xlsx"
#     render = xlrender.xlrender(template, deffile)
#     data = {
#         "data": []
#     }

#     headerobj = {
#         "chuquan": "HANOI",
#         "tendonvi": "ABC",
#         #"ngaybaocao": u"..., ngày " + "" + " tháng " + "" + " năm " + str(baocao.nambaocao)
#         "ngaybaocao": '....., ngày %d tháng %d năm %d' % (1,2,2020),
#         "tungaydenngay": ""
#     }
    
#     render.put_area("header", ujson.dumps( headerobj))
#     render.save(filename)
    
#     return await file(filename, mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#         filename='ds-cuakhau.xlsx')



notdict = ['_created_at','_updated_at','_deleted','_deleted_at','_etag', 'id', 'baocao_id', 'cuakhau_id', 'sothutu', 'stt']

def getcongdon(baocaos, attrname):
    congdon = {}
    for bc in baocaos:
        for obj in getattr(bc, attrname).all():
            if obj.tencuakhau not in congdon:
                congdon[obj.tencuakhau] = {}
            for col in obj.__table__.c:
                if (col.name not in notdict) and (str(col.type) in ['INTEGER','SMALLINT', 'FLOAT']):
                    congdon[obj.tencuakhau][col.name] = (congdon[obj.tencuakhau][col.name] + (getattr(obj, col.name) if getattr(obj, col.name) is not None else 0)) \
                    if (col.name in congdon[obj.tencuakhau]) and (congdon[obj.tencuakhau][col.name] is not None) else getattr(obj, col.name)
    return congdon

def processBangCongDon(render, tenbang, congdon, baocao):
    bangdata = getattr(baocao, "bang" + tenbang)
    tongso = {}
    render.put_area(tenbang + "header")
    #render.put_area(range)
   
    for chitiet in bangdata:
        chitietobj = to_dict(chitiet)
        #tinhtoan cong don
        #for col in obj.__table__.c:
        if chitiet.tencuakhau in congdon:
            for key, value in (congdon[chitiet.tencuakhau]).items():
                chitietobj[key] = chitietobj[key] if chitietobj[key] is not None else 0
                chitietobj['cd_'+ key] = chitietobj[key] + (value if value is not None else 0)
                
                tongso[key] = tongso[key] + chitietobj[key] if key in tongso else chitietobj[key]
                tongso['cd_'+ key] = tongso['cd_'+ key] + chitietobj['cd_'+ key] if 'cd_'+ key in tongso else chitietobj['cd_'+ key]
        else:
            for col in chitiet.__table__.c:
                key = col.name
                if (key not in notdict) and (str(col.type) in ['INTEGER','SMALLINT', 'FLOAT']):
                    chitietobj[key] = chitietobj[key] if chitietobj[key] is not None else 0
                    chitietobj['cd_'+ key] = chitietobj[key]
                    tongso[key] = tongso[key] + chitietobj[key] if key in tongso else chitietobj[key]
                    tongso['cd_'+ key] = tongso['cd_'+ key] + chitietobj['cd_'+ key] if 'cd_'+ key in tongso else chitietobj['cd_'+ key]
                    
        #end congdong
        for key, value in chitietobj.items():
            if chitietobj[key] == 0:
                chitietobj[key] = None
            
        render.put_area(tenbang + "chitiet", ujson.dumps(chitietobj))
        #render.put_area(range)
    
    for key, value in tongso.items():
        if tongso[key] == 0:
            tongso[key] = None
      
    render.put_area(tenbang + "tongso", ujson.dumps(tongso))
    #render.put_area(range)
    
    render.put_area(tenbang + "footer")
    
    if baocao.loaikybaocao > 2:
        render.put_area("nhanxetkiennghi", ujson.dumps({"nhanxet":getattr(baocao, "nhanxet" + tenbang)}))
    
def processBang(render, tenbang, baocao):
    bangdata = getattr(baocao, "bang" + tenbang)
    
    render.put_area(tenbang + "header")
    #render.put_area(range)
    
    for chitiet in bangdata:
        chitietobj = to_dict(chitiet)
        render.put_area(tenbang + "chitiet", ujson.dumps(chitietobj))
        #render.put_area(range)
    render.put_area(tenbang + "footer")
    #render.put_area(range)
    
def processBangKinhPhi(render, baocao):
    bangdata = getattr(baocao, "bangkinhphi")
    
    render.put_area("kinhphiheader")
    #render.put_area(range)
    
    tongso = 0
    for chitiet in bangdata:
        chitietobj = to_dict(chitiet)
        tongso = tongso + (chitiet.kinhphi if chitiet.kinhphi is not None else 0)
        render.put_area("kinhphichitiet", ujson.dumps(chitietobj))
        #render.put_area(range)
        
    render.put_area("kinhphitongso", ujson.dumps({"tongso":tongso}))
    #render.put_area(range)  
    render.put_area("kinhphifooter")
    #render.put_area(range)
    
def processCuaKhau(render, cuakhaulist):
    render.put_area("cuakhauheader")
    #render.put_area(range)
    #chitiet
    for chitiet in cuakhaulist:
        chitietobj = to_dict(chitiet)
        render.put_area("cuakhauchitiet", ujson.dumps(chitietobj))
        #render.put_area(range)
        
    render.put_area("cuakhaufooter")
    #render.put_area(range)
    
@exportbp.route('/excel/baocao/<id>')
async def exportbaocao(request, id=None):
    starttime = time.time()
    #print("export baocao")
    if (id is not None):
        baocao = BaoCao.query.filter(BaoCao.id == int(id)).first()
        if baocao is not None:
            deffile = None
            template = None
            
            baocaos = BaoCao.query.filter(BaoCao.donvi_id == baocao.donvi_id).filter(BaoCao.loaikybaocao == baocao.loaikybaocao).filter(BaoCao.loaibaocao == baocao.loaibaocao).\
                            filter(BaoCao.nambaocao == baocao.nambaocao).filter(BaoCao.kybaocao < baocao.kybaocao)
            
            if baocao.loaibaocao == 2:
                baocaos = baocaos.filter(BaoCao.cuakhau_id == baocao.cuakhau_id)
            baocaos = baocaos.all()
                
            congdon = {
                "bangnguoi": getcongdon(baocaos, 'bangnguoi') ,
                "bangphuongtien": getcongdon(baocaos, 'bangphuongtien') ,
                "banghanghoa": getcongdon(baocaos, 'banghanghoa') ,
                "bangthithe": getcongdon(baocaos, 'bangthithe') ,
                "bangvssh": getcongdon(baocaos, 'bangvssh') ,
            }
            
            #print congdon
            
            if baocao.loaikybaocao == 1:
                deffile = "exceltpl/baocaotuan.json"
                template = "exceltpl/baocaotuan.xlsx"
            if baocao.loaikybaocao == 2:
                deffile = "exceltpl/baocaothang.json"
                template = "exceltpl/baocaothang.xlsx"
            if baocao.loaikybaocao == 3:
                deffile = "exceltpl/baocao6thang.json"
                template = "exceltpl/baocao6thang.xlsx"
            if baocao.loaikybaocao == 4:
                deffile = "exceltpl/baocao6thang.json"
                template = "exceltpl/baocao9thang.xlsx"
            if baocao.loaikybaocao == 5:
                deffile = "exceltpl/baocaonam.json"
                template = "exceltpl/baocaonam.xlsx"
            
            render = xlrender.xlrender(template, deffile)
            
            tendonvi = ""
            chuquan = ""
            if baocao.loaibaocao == 1:
                tendonvi = baocao.donvi.ten if baocao.donvi is not None else ""
                chuquan = baocao.donvi.coquanchuquan if baocao.donvi is not None else ""
            if baocao.loaibaocao == 2:
                tendonvi = baocao.cuakhau.ten if baocao.cuakhau is not None else ""
                chuquan = baocao.donvi.ten if baocao.donvi is not None else ""
                    
            headerobj = {
                "chuquan": chuquan,
                "tendonvi": tendonvi,
                "ma": u"Số: " + (baocao.ma if baocao.ma is not None else ""),
                #"ngaybaocao": u"..., ngày " + "" + " tháng " + "" + " năm " + str(baocao.nambaocao)
                "ngaybaocao": u'....., ngày %d tháng %d năm %d' % (baocao.ngaybaocao.day,baocao.ngaybaocao.month,baocao.ngaybaocao.year),
                "kybaocao": baocao.kybaocao,
                "tungaydenngay": u'(Từ ngày %d tháng %d năm %d đến ngày %d tháng %d năm %d)' % (baocao.tungay.day,baocao.tungay.month,baocao.tungay.year,\
                                                                                                baocao.denngay.day,baocao.denngay.month,baocao.denngay.year)
            }
            
            render.put_area("header", ujson.dumps( headerobj))
            #render.put_area(range)
            
            if baocao.loaikybaocao > 4:
                cuakhaus = None
                if(baocao.donvi is not None):
                    cuakhaus = baocao.donvi.cuakhau.all()
                    processCuaKhau(render, cuakhaus)
                
                #nhan luc
                render.put_area("nhanluc", ujson.dumps({"nhanlucts": baocao.nhanlucts,
                                                      "nhanlucbienche": baocao.nhanlucbienche,
                                                      "nhanluchopdong": baocao.nhanluchopdong
                                                      }))
                #render.put_area(range)
                
                processBang(render, 'trinhdochuyenmon', baocao)
                processBang(render, 'daotaocanbo', baocao)
                
            #nguoi
            
            processBangCongDon(render, 'nguoi', congdon['bangnguoi'], baocao)
            processBangCongDon(render, 'phuongtien', congdon['bangphuongtien'], baocao)
            
            if baocao.loaikybaocao > 1:
                processBangCongDon(render, 'hanghoa', congdon['banghanghoa'], baocao)
                processBangCongDon(render, 'thithe', congdon['bangthithe'], baocao)
                processBangCongDon(render, 'vssh', congdon['bangvssh'], baocao)
                
            #bangkhac
            if baocao.loaikybaocao > 2:
                processBang(render, 'tiemchung', baocao)
                processBang(render, 'dichbenh', baocao)
                processBang(render, 'xetnghiem', baocao)
                
                if baocao.loaikybaocao > 4:
                    processBang(render, 'truyenthong', baocao)
                    processBang(render, 'nghiencuukhoahoc', baocao)
                    processBang(render, 'hoptacquocte', baocao)
                    
                    render.put_area("hoatdongkhac", ujson.dumps({"hoatdongkhac": baocao.hoatdongkhac}))
                    #render.put_area(range)
                    
                processBangKinhPhi(render, baocao)
                
            footerobj = {
                "hoatdongkhac": baocao.hoatdongkhac,
                "nhanxet": baocao.nhanxet,
                "kiennghi": baocao.kiennghi
            }
            
            render.put_area("footer", ujson.dumps(footerobj))
            #render.put_area(range)
         
            #save:
            #out = StringIO.StringIO()
            #render.save(out)
            #out.seek(0)
            filename = "output-baocao" + str(id)+".xlsx"
            render.save(filename)
            
            endtime = time.time()
            print(endtime - starttime)
            
            return await file(filename, mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename='baocao-' + str(id) + '.xlsx')

    return text("Không tìm thấy báo cáo...", status=404)

def processThongKeBang(render, bang, bangdata, solieucuakhau):
    
    tongso = {"donvi_ten": u"Tổng số"}
    sttdonvi = 0
    if(bangdata is not None):
        for donvi, donvidata in bangdata.items():
            sttdonvi = sttdonvi + 1
            donviobj = {"sothutu": sttdonvi}
            
            for key, val in donvidata.items():
                if(key != "cuakhau"):
                    donviobj[key] = val
                    if(key != "donvi_ten"):
                        tongso[key] = tongso[key] + (val if val is not None else 0) if ((key in tongso) and (tongso[key] is not None)) else (val if val is not None else 0)
                
            render.put_area(bang + "chitiet", ujson.dumps(donviobj))    
            
                    #val == cuakhau
            if (donvidata is not None) and ("cuakhau" in donvidata):
                sttcuakhau = 0
                if solieucuakhau is True:
                    for cuakhaudata in donvidata["cuakhau"]:
                        sttcuakhau = sttcuakhau + 1
                        cuakhaudata["sothutu"] = sttcuakhau
                        render.put_area(bang + "ckchitiet", ujson.dumps(cuakhaudata))
                            #render.put_area("bangnguoickchitiet")
            #print donviobj
    render.put_area(bang + "chitiet", ujson.dumps(tongso))
    render.put_area(bang + "footer")
    

def exportthongketheodonvi(data, filename):
    if(data is not None):
        deffile = "exceltpl/thongketheodonvi.json"
        template = "exceltpl/thongketheodonvi.xlsx"
        render = xlrender.xlrender(template, deffile)
        
        ngaybaocao = datetime.now()
        tungaydenngay = None
        solieucuakhau = False
        fmt = "%Y-%m-%d"
        if("param" in data):
            #tungaydate =  datetime.strptime(data["param"]["tungay"], fmt) 
            #denngaydate =  datetime.strptime(data["param"]["denngay"], fmt) 
            #tungaydenngay = u'(Từ ngày %d tháng %d năm %d đến ngày %d tháng %d năm %d)' % (tungaydate.day,tungaydate.month,tungaydate.year, denngaydate.day,denngaydate.month,denngaydate.year)
            if data["param"]["loaikybaocao"] == 1:
                tungaydenngay = u'(Từ tuần %d năm %d đến tuần %d năm %d)' % (data["param"]["tutuan"],data["param"]["tunam"],data["param"]["dentuan"], data["param"]["dennam"])
            if data["param"]["loaikybaocao"] == 2:
                tungaydenngay = u'(Từ tháng %d năm %d đến tháng %d năm %d)' % (data["param"]["tuthang"],data["param"]["tunam"],data["param"]["denthang"], data["param"]["dennam"])
            if data["param"]["loaikybaocao"] == 3:
                tungaydenngay = u'(Từ 6 tháng kỳ %d năm %d đến 6 tháng kỳ %d năm %d)' % (data["param"]["tu6thang"],data["param"]["tunam"],data["param"]["den6thang"], data["param"]["dennam"])
            if data["param"]["loaikybaocao"] > 3:
                tungaydenngay = u'(Từ năm %d đến năm %d)' % (data["param"]["tunam"], data["param"]["dennam"])
            
            solieucuakhau = data["param"]["solieucuakhau"] 
        headerobj = {
                "chuquan": current_user.donvi.coquanchuquan,
                "tendonvi": current_user.donvi.ten,
                #"ngaybaocao": u"..., ngày " + "" + " tháng " + "" + " năm " + str(baocao.nambaocao)
                "ngaybaocao": u'....., ngày %d tháng %d năm %d' % (ngaybaocao.day,ngaybaocao.month,ngaybaocao.year),
                "tungaydenngay": tungaydenngay
            }
            
        render.put_area("header", ujson.dumps( headerobj))
        
        bangs = ["bangnguoi", "bangphuongtien", "banghanghoa", "bangthithe", "bangvssh"]
        
        #for bang, bangdata in data.items():
        if "data" in data:
            for bang in bangs:
                if bang in data["data"]:
                    render.put_area((bang + "header"))
                    processThongKeBang(render, bang, data["data"][bang], solieucuakhau)
            
        render.save(filename)
    return

@exportbp.route('/excel/cuakhau')
async def exportcuakhau(request):
    filter = None
    filterstr = request.args.get("filter", None)
    filter = json.loads(filterstr) if filterstr is not None else None
    
    cuakhaus = CuaKhau.query
    
    if (filter is not None) and (filter["donvi_id"] is not None):
        #cuakhaus = cuakhaus.filter(CuaKhau.donvi_id == filter["donvi_id"])
        
        cuakhaus = cuakhaus.filter(CuaKhau.donvi_id.in_(filter["donvi_id"]))
        
    if (filter is not None) and (filter["tinhthanh_id"] is not None):
        cuakhaus = cuakhaus.filter(CuaKhau.tinhthanh_id == filter["tinhthanh_id"])
    
    if (filter is not None) and (filter["ten"] is not None):
        cuakhaus = cuakhaus.filter(CuaKhau.ten.contains(filter["tinhthanh_id"]))
        
    if (filter is not None) and (filter["loaicuakhau"] is not None) and (len(filter["loaicuakhau"]) > 0):
        orfilter = None
        
        if 1 in filter["loaicuakhau"]:
            #cuakhaus = cuakhaus.filter(CuaKhau.duongboquocte == True)
            orfilter = orfilter | (CuaKhau.duongboquocte == True) if orfilter is not None else (CuaKhau.duongboquocte == True)
        if 2 in filter["loaicuakhau"]:
            #cuakhaus = cuakhaus.filter(CuaKhau.duongbochinh == True)
            orfilter = orfilter | (CuaKhau.duongbochinh == True) if orfilter is not None else (CuaKhau.duongbochinh == True)
        if 3 in filter["loaicuakhau"]:
            #cuakhaus = cuakhaus.filter(CuaKhau.duongbophu == True)
            orfilter = orfilter | (CuaKhau.duongbophu == True) if orfilter is not None else (CuaKhau.duongbophu == True)
        if 4 in filter["loaicuakhau"]:
            #cuakhaus = cuakhaus.filter(CuaKhau.duongsat == True)
            orfilter = orfilter | (CuaKhau.duongsat == True) if orfilter is not None else (CuaKhau.duongsat == True)
        
        cuakhaus = cuakhaus.filter(orfilter)
        
    donvichildids =  current_user.donvi.children_ids()
    
    cuakhaus = cuakhaus.filter(CuaKhau.donvi_id.in_(donvichildids))
    
    cuakhaus = cuakhaus.order_by(CuaKhau.donvi_id.asc()).all()
    
    deffile = "exceltpl/cuakhau_template.json"
    template = "exceltpl/cuakhau_template.xlsx"
    render = xlrender.xlrender(template, deffile)
    
    render.put_area("header")
    for obj in cuakhaus:
        data = {
            "ma": obj.ma,
            "ten": obj.ten,
            "diachi": obj.diachi,
            "tinhthanh": obj.tinhthanh.ten if obj.tinhthanh is not None else "",
            "donvi": obj.donvi.ten if obj.donvi is not None else "",
            "nguoilienlac": obj.nguoilienlac,
            "sodienthoai": obj.sodienthoai,
            "email":obj.email,
            "duongboquocte": "X" if obj.duongboquocte else "",
            "duongbochinh": "X" if obj.duongbochinh else "",
            "duongbophu": "X" if obj.duongbophu else "",
            "duongsat": "X" if obj.duongsat else "",
            "duonghangkhong": "X" if obj.duonghangkhong else "",
            "duongthuyloai1": "X" if obj.duongthuyloai1 else "",
            "duongthuyloai2": "X" if obj.duongthuyloai2 else "",
        }
        render.put_area("chitiet", ujson.dumps(data))
    render.put_area("footer")
    
    filename = "output-ds-cuakhau.xlsx"
    render.save(filename)

    return await file(filename, mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename='ds-cuakhau-' + str(time.time())+'.xlsx')

list_char = [
    'A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W', 'X', 'Y', 'Z',
    'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AR', 'AS', 'AT', 'AU', 'AV', 'AW', 'AX', 'AY', 'AZ',
    'BA', 'BB', 'BC', 'BD', 'BE', 'BF', 'BG', 'BH', 'BI', 'BJ', 'BK', 'BL', 'BM', 'BN', 'BO', 'BP', 'BQ', 'BR', 'BS', 'BT', 'BU', 'BV', 'BW', 'BX', 'BY', 'BZ'
]

center_alignment = Alignment(horizontal="center", vertical="center")
bold_font = Font(bold=True)
black_side = Side(border_style="thin",color='00000000')
black_border = Border(left=black_side, right=black_side, top=black_side, bottom=black_side)

def processChiTieu(ws, chitieu, idx,  data, start_row):
    len_day = len(data["days"])
    len_col = len(data["ten"])
    tongso_col = list_char[3+len_col]
    ws.merge_cells('A' + str(start_row) +':A' + str(len_day + start_row))
    ws.merge_cells('B' + str(start_row) +':B' + str(len_day + start_row - 1))

    

    ws['A' + str(start_row)] =  str(idx + 1)
    ws['B' + str(start_row)] =  chitieu["text"]
    ws['A' + str(start_row)].alignment = center_alignment
    ws['B' + str(start_row)].alignment = center_alignment
    
    # ws['A' + str(start_row)].font = bold_font
    # ws['B' + str(start_row)].font = bold_font

    ws.merge_cells('B' + str(start_row + len_day) +':C' + str(len_day + start_row))
    ws['B' + str(start_row + len_day)] = 'Cộng dồn'
    ws['B' + str(start_row + len_day)].font = bold_font
    ws['B' + str(start_row + len_day)].alignment = center_alignment

    #chitiet data
    for i in range(len(data["days"])):
        tongso = 0;
        ws['C' + str(start_row + i)] = data["days"][i]
        for j in range(len_col):
            ckten = data["ten"][j]
            col = list_char[3+j]
            value = data["data"][j]["data_value"][i][chitieu["name"]]
            if (value is not None):
                ws[col + str(start_row + i) ] = data["data"][j]["data_value"][i][chitieu["name"]]
                tongso = tongso + value
        
        
        if tongso > 0:
            ws[tongso_col + str(start_row + i)] = tongso
        ws[tongso_col + str(start_row + i)].font = bold_font


    tong_tongso = 0
    for j in range(len_col):
        tongdoc = 0
        col = list_char[3+j]
        for i in range(len_day):
            value = data["data"][j]["data_value"][i][chitieu["name"]]
            if (value is not None):
                tongdoc = tongdoc + value
                tong_tongso = tong_tongso + value
        if tongdoc > 0:
            ws[col + str(start_row + len_day) ] = tongdoc
        ws[col + str(start_row + len_day) ].font = bold_font

    ws[tongso_col + str(start_row + len_day) ] = tong_tongso
    ws[tongso_col + str(start_row + len_day) ].font = bold_font
    for i in range(start_row, start_row + len_day + 1):
        for j in range(len_col + 4):
            ws[list_char[j] + str(i)].border = black_border
            ws[list_char[j] + str(i)].alignment = center_alignment

    return start_row + len_day + 1


#@exportbp.route('/excel/cuakhau')
async def exportthongkenghingonhiembenhnhoma(request, data):
    from .nghingonhiembenhnhoma import chitieu_nghingonhiembenhnhoma as chitieus
    filename=data.get("filename")
    title = "cửa khẩu"
    if(data.get("tuyen") == "donvi"):
        title = "đơn vị"
    
    # chitieus = [
	# 			{"name":"hanhkhach", "text": "Số lượt khách khai báo y tế"},
	# 			{"name":"chuyenbay", "text": "Số lượng chuyến bay nhập"},
	# 			{"name":"nguoinghingo", "text": "Trường hợp nghi ngờ mắc bệnh truyền nhiễm"}
	# 		]
    wb = Workbook()
    ws = wb.active

    ws.merge_cells('A2:D2')
    ws.merge_cells('A3:D3')
    ws.merge_cells('A4:K4')

    

    ws['A2'] = "Cục Y tế Dự phòng"
    ws['A3'] = data.get("tendonvi")
    ws['A3'].font = bold_font
    ws['A4'] = "BÁO CÁO SỐ LIỆU LƯỢT KHÁCH KHAI BÁO Y TẾ TẠI CÁC " + title.upper()+" NĂM 2020"
    ws['A4'].font = bold_font

    #table data
    ws.merge_cells('A6:A7')
    ws.merge_cells('B6:B7')
    ws.merge_cells('C6:C7')

    ws['A6'] = 'STT'
    ws['A6'].font = bold_font
    ws['B6'] = 'Nội dung báo cáo'
    ws['B6'].font = bold_font
    ws['C6'] = 'Ngày tháng'
    ws['C6'].font = bold_font

    ws['A2'].alignment = center_alignment
    ws['A3'].alignment = center_alignment
    ws['A4'].alignment = center_alignment
    ws['A6'].alignment = center_alignment
    ws['B6'].alignment = center_alignment
    ws['C6'].alignment = center_alignment


    len_col = len(data["ten"])
    max_col = list_char[3+len_col-1]
    tongso_col = list_char[3+len_col]
    #print(max_col)
    ws.merge_cells('D6:'+max_col+'6')
    ws.merge_cells(tongso_col + '6:'+tongso_col+'7')
    ws[tongso_col + '6'] = 'Tổng số'
    ws[tongso_col + '6'].font = bold_font
    ws[tongso_col + '6'].alignment = center_alignment

    ws['D6'] = 'Tên ' + title
    ws['D6'].alignment = center_alignment
    ws['D6'].font = bold_font

    for i in range(len(data["ten"])):
        tencuakhau = data["ten"][i]
        col = list_char[3+i]
        ws[col + '7'] = tencuakhau
        ws[col + '7'].font = bold_font
    
    start_row = 6
    for i in range(start_row, 8):
        for j in range(len_col + 4):
            ws[list_char[j] + str(i)].border = black_border
        

    #het header
    start_row = 8
    for i in range(len(chitieus)):
        start_row = processChiTieu(ws, chitieus[i], i, data, start_row)
     


    wb.save(filename = filename)
    return await file(filename, mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename=filename)


    return json(data)


@exportbp.route('/excel/truonghopcachlytaptrung')
async def exporttruonghopcachlytaptrung(request):
    donvi_id = request.args.get("donvi_id")
    
    if donvi_id is None:
        return text("Không tìm thấy báo cáo...", status=404)
    filename  = "truonghopcachlytaptrung-" + donvi_id + "-" + str(time.time())+'.xlsx'
    donvi = db.session.query(DonVi).filter(DonVi.id == int(donvi_id)).first()

    cuakhau_id = request.args.get("cuakhau_id")
    ngaybaocao = request.args.get("ngaybaocao", "")

    items = TruongHopCachLyTapTrung.query.filter(TruongHopCachLyTapTrung.donvi_id == int(donvi_id))
    if cuakhau_id is not None:
        items = items.filter(TruongHopCachLyTapTrung.cuakhau_id == int(cuakhau_id))
    if ngaybaocao is not None:
        items = items.filter(TruongHopCachLyTapTrung.ngaybaocao == ngaybaocao)
    items = items.all()

    wb = Workbook()
    ws = wb.active

    ws.merge_cells('A2:D2')
    ws.merge_cells('A3:D3')
    ws.merge_cells('D4:K4')

    

    ws['A2'] = "Cục Y tế Dự phòng"
    ws['A3'] = donvi.ten
    ws['A3'].font = bold_font
    ws['D4'] = "DANH SÁCH THEO DÕI KHÁCH NHẬP CẢNH ĐƯỢC CÁCH LY"
    ws['D4'].font = bold_font
    ws['B5'] = "Ngày: "

    if ngaybaocao is not None:
        ngaybaocao_date = datetime.strptime(ngaybaocao, "%Y-%m-%dT%H:%M:%S")
        print(ngaybaocao_date)
        ws['C5'] = ngaybaocao_date.strftime("%d/%m/%Y")

    ws['A7'] = "STT"
    ws['B7'] = "Họ và tên"
    ws['C7'] = "CMTND/Hộ chiếu"
    ws['D7'] = "Tuổi"
    ws['E7'] = "Giới tính"
    ws['F7'] = "Nghề nghiệp"
    ws['G7'] = "Xã"
    ws['H7'] = "Huyện"
    ws['I7'] = "Tỉnh"
    ws['J7'] = "Quốc tịch"
    ws['K7'] = "Số điện thoại"
    ws['L7'] = "Tiền sử dịch tễ"
    ws['M7'] = "Triệu chứng"
    ws['N7'] = "Địa điểm cách ly"
    ws['O7'] = "Bệnh lý nền kèm theo"
    ws['P7'] = "Ghi chú"
    ws['Q7'] = "Nhập cảnh tại của khẩu"

    for j in range(17):
        ws[list_char[j] + "7"].alignment = center_alignment
        ws[list_char[j] + "7"].border = black_border
        ws[list_char[j] + "7"].font = bold_font

    start_row = 8

    for item in items:
        obj = to_dict(item)
        ws['A' + str(start_row)] = str(start_row - 7)
        ws['B' + str(start_row)] = obj["hoten"]
        ws['C' + str(start_row)] = obj["cmtnd"]
        ws['D' + str(start_row)] = obj["namsinh"]
        ws['E' + str(start_row)] = obj["gioitinh"]
        ws['F' + str(start_row)] = obj["nghenghiep"]
        ws['G' + str(start_row)] = obj["noio_xaphuong"]
        ws['H' + str(start_row)] = obj["noio_quanhuyen"]
        ws['I' + str(start_row)] = obj["noio_tinhthanh"]
        ws['J' + str(start_row)] = obj["quoctich"]
        ws['K' + str(start_row)] = obj["sodienthoai"]
        ws['L' + str(start_row)] = obj["tiensu_dichte"]
        ws['M' + str(start_row)] = obj["tiensu_trieuchunglamsang"]
        ws['N' + str(start_row)] = obj["noitiepnhan_xutri"]
        ws['O' + str(start_row)] = obj["benhly_kemtheo"]
        ws['P' + str(start_row)] = obj["ghichu"]
        ws['Q' + str(start_row)] = obj["tencuakhau"]
        start_row = start_row + 1

    for i in range(8, start_row + 1):
        for j in range(17):
            ws[list_char[j] + str(i)].border = black_border
            ws[list_char[j] + str(i)].alignment = center_alignment


    wb.save(filename = filename)
    return await file(filename, mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename=filename)

    return json({})
